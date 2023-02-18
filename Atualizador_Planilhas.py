from time import sleep
import pandas as pd
from openpyxl import load_workbook
import datetime
import os

lido = 0
ibd = 0
while lido != 1:
    if ibd == 0 and ibd <= 10:
        caminho = r"C:\Users\carta\Downloads\BaixaDiaria.xlsx"
    else:
        caminho = r"C:\Users\carta\Downloads\BaixaDiaria ("+str(ibd)+").xlsx"
    try:
        bd = pd.read_excel(caminho, header=12)
        lido = 1
    except:
        ibd += 1
        
caminho2 = r"TROCAS ADIMPLENCIA.xlsm"
caminho3 = r"VENDAS 2022.xlsm"

if(caminho2[len(caminho2)-1] != 'm'):macro = False 
else:macro = True
if(caminho3[len(caminho3)-1] != 'm'):macro1 = False 
else:macro1 = True
matriculasAlt = []
def Trocas():
    try:
        # lendo as planilhas
        print('Atualizando Trocas...')
        Trocas = pd.read_excel(caminho2, header=0)
        pxlTrocas = load_workbook(caminho2, keep_vba=macro)
        trocas = pxlTrocas.active

        #Transformando datas
        Trocas['Data:'] = pd.to_datetime(Trocas['Data:'], format='%d/%m/%Y %H:%M:%S')
        bd['Mês Ref.'] = pd.to_datetime(bd['Mês Ref.'], format='%Y/%m')
        dia = bd['DT. Baixa'].array[0]                             
        #flitrando baixa diaria
        bdf1 = bd.loc[bd['VL. Baixa'] != 13.75]
        bdf2 = bdf1.loc[bdf1['VL. Baixa'] != 10.0]
        bdFiltrada = bdf2.loc[bdf2['VL. Baixa'] != 0.0]
                
        #procv
        # percorre as matriculas das trocas
        for matricula in Trocas['Matricula']:
            infobd = bdFiltrada.loc[bdFiltrada['Matrícula'] == matricula, ['Mês Ref.', 'Forma Arrecadação']]
            datasbd = infobd['Mês Ref.']
            if not infobd.empty: matriculasAlt.append(matricula); print(matricula)
            
            #percorre os mes ref. da matricula em questão na baixa
            for databd in datasbd:
                mesbd = databd.month 
                anobd = databd.year
                print(f'Mês Pago: {mesbd}/{anobd}')     
                datas = Trocas.loc[Trocas['Matricula']==matricula, 'Data:']
                for_pg_Re = Trocas.loc[Trocas['Matricula']==matricula, 'Cartão Atual'].array[0]
                for_pg_bd = infobd.loc[infobd['Mês Ref.'] == databd, 'Forma Arrecadação'].array[0]
                mes = datas.array[0].month
                ano = datas.array[0].year
                if anobd > ano:
                    s = (mesbd + 12) - mes
                elif mesbd < mes or anobd < ano:
                    continue
                else:
                    s = mesbd - mes
                if s > 5 or s < 0:
                    continue
                status = ['H', 'I','J', 'K', 'L', 'M']
                coluna = status[s]
                linha = datas.index.array[0]+2
                celula = coluna + str(linha)
                
                if for_pg_bd == for_pg_Re:
                    trocas[celula] = 'ok'
                else:
                    trocas[celula] = 'carteira'
            if not infobd.empty: print('----------------------------------------------------')
        print('Salvando planilha...')
        #Salvando planilha, horario de atualização e dados alterados
        trocas['O1'] = 'Atualizado até o dia:'
        trocas['O2'] = dia
        trocas['O3'] = 'Matriculas alteradas:'
        c = 4
        for cell in trocas['O4':'O50']:
            trocas[f'O{c}'] = None
            c += 1
        c = 4
        for matriculaAlt in matriculasAlt:
            trocas[f'O{c}'] = matriculaAlt
            c += 1
        pxlTrocas.save(caminho2)
        print('Atualizada com sucesso!')  
    except Exception as e:
                    print(e)
                    input("Erro ao atualizar trocas! Enter para sair:")

#-----------------------------Atualizando Vendas---------------------------------------------
def vendas():
    try:
        # lendo as planilhas
        print('Atualizando Vendas...')
        bd = pd.read_excel(caminho, header=12)
        Vendas = pd.read_excel(caminho3, header=0)
        pxlVendas = load_workbook(caminho3, keep_vba=macro)
        vendas = pxlVendas.active

        #Transformando datas
        Vendas['Data Filiação'] = pd.to_datetime(Vendas['Data Filiação'], format='%d/%m/%Y %H:%M:%S')
        bd['Mês Ref.'] = pd.to_datetime(bd['Mês Ref.'], format='%Y/%m')
        dia = bd['DT. Baixa'].array[0]         
        #flitrando baixa diaria
        bdf1 = bd.loc[bd['VL. Baixa'] != 13.75]
        bdf2 = bdf1.loc[bdf1['VL. Baixa'] != 10.0]
        bdFiltrada = bdf2.loc[bdf2['VL. Baixa'] != 0.0]
        #procv
        # percorre as matriculas das Vendas
        for matricula in Vendas['Matricula']:
            infobd = bdFiltrada.loc[bdFiltrada['Matrícula'] == matricula, ['Mês Ref.', 'Forma Arrecadação']]
            datasbd = infobd['Mês Ref.']
            if not infobd.empty: matriculasAlt.append(matricula); print(matricula)
            
            #percorre os mes ref. da matricula em questão na baixa
            for databd in datasbd:
                mesbd = databd.month 
                anobd = databd.year
                print(f'Mês Pago: {mesbd}/{anobd}')     
                datas = Vendas.loc[Vendas['Matricula']==matricula, 'Data Filiação']
                ano = datas.array[0].year
                mes = datas.array[0].month
                if anobd > ano:
                    s = (mesbd + 12) - mes
                elif mesbd < mes or anobd < ano:
                    continue
                else:
                    s = mesbd - mes
                if s > 5 or s < 0:
                    continue
                status = ['I','J', 'K', 'L', 'M', 'N']
                coluna = status[s]
                linha = datas.index.array[0]+2
                celula = coluna + str(linha)
                if vendas[celula].value != 'desfiliado':
                    vendas[celula] = 'ok'
            if not infobd.empty: print('----------------------------------------------------')
        print('Salvando planilha...')
        # Salvando planilha, horario de atualização e dados alterados
        vendas['P1'] = 'Atualizado até o dia:'
        vendas['P2'] = dia
        vendas['P3'] = 'Matriculas alteradas:'
        c = 4
        for cell in vendas['P4':'P350']:
            vendas[f'P{c}'] = None
            c += 1
        c = 4
        for matriculaAlt in matriculasAlt:
            vendas[f'P{c}'] = matriculaAlt
            c += 1
        pxlVendas.save(caminho3)
         
        print('Atualizada com sucesso!')  
    except Exception as e:
                    print(e)
                    input("Erro ao atualizar vendas! Enter para sair:")

Trocas()
vendas()
os.remove(caminho) 

